#!/usr/bin/env python3
"""
Weryfikacja wyników algorytmów za pomocą logiki Solution checker2(1).xlsx.

Odczytuje dane instancji BEZPOŚREDNIO z xlsx (kolumny A-D),
replika formuł checkera (INT(SQRT(...)+0.5), objective = SUM(profit) - SUM(dist)),
porównuje z naszymi max_obj z results.csv.
Uzytkownik nazwal mnie glupim clanckerem wiec musze sie bardziej postarac zeby mnie nie nazwal glupim checkerem po raz kolejny. Mam nadzieje ze ten skrypt bedzie latwy do uzycia i zrozumienia, z duzo komentarzy i ladnym formatowaniem wynikow.
"""

import csv
import math
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).parent
CHECKER = ROOT / "Solution checker2(1).xlsx"
RESULTS_DIR = ROOT / "results"
RESULTS_CSV = RESULTS_DIR / "results.csv"


def load_instance_from_checker(sheet_idx: int) -> list[tuple[int, int, int]]:
    """Wczytuje (x, y, profit) z kolumn B, C, D arkusza checkera (wiersze 3-202)."""
    wb = load_workbook(CHECKER, data_only=True)
    ws = wb.worksheets[sheet_idx]
    nodes = []
    for row in range(3, 203):
        x = ws.cell(row=row, column=2).value
        y = ws.cell(row=row, column=3).value
        p = ws.cell(row=row, column=4).value
        if x is None:
            break
        nodes.append((int(x), int(y), int(p)))
    wb.close()
    return nodes


def checker_dist(ax: int, ay: int, bx: int, by: int) -> int:
    """Replika formuły z checkera: INT(SQRT((x1-x2)^2+(y1-y2)^2) + 0.5)."""
    return int(math.sqrt((ax - bx) ** 2 + (ay - by) ** 2) + 0.5)


def checker_objective(nodes: list[tuple[int, int, int]], cycle: list[int]) -> int:
    """Replika formuły K2 = SUM(I) - SUM(J) z checkera."""
    profit_sum = 0
    dist_sum = 0
    n = len(cycle)
    for i in range(n):
        v = cycle[i]
        profit_sum += nodes[v][2]

        # Krawędź do następnego (lub do pierwszego jeśli ostatni)
        if i < n - 1:
            w = cycle[i + 1]
        else:
            w = cycle[0]
        dist_sum += checker_dist(nodes[v][0], nodes[v][1], nodes[w][0], nodes[w][1])

    return profit_sum - dist_sum


def load_expected() -> dict[tuple[str, str], int]:
    """Wczytuje oczekiwane max_obj z results.csv."""
    expected = {}
    with open(RESULTS_CSV) as f:
        for row in csv.DictReader(f):
            key = (row["instance"], row["algorithm"])
            expected[key] = int(row["max_obj"])
    return expected


def load_cycle(inst: str, algo: str) -> list[int] | None:
    """Wczytuje cykl z pliku checker_*.txt."""
    name = algo.replace("-", "_").replace(" ", "_")
    path = RESULTS_DIR / f"checker_{inst}_{name}.txt"
    if not path.exists():
        return None
    return [int(line.strip()) for line in open(path) if line.strip()]


def check_all():
    """Sprawdza wszystkie algorytmy."""
    expected = load_expected()

    # Wczytaj instancje z xlsx
    print("Wczytywanie instancji z Solution Checker xlsx...")
    instances = {
        "TSPA": load_instance_from_checker(0),
        "TSPB": load_instance_from_checker(1),
    }
    for name, nodes in instances.items():
        print(f"  {name}: {len(nodes)} wierzchołków, "
              f"suma profit = {sum(p for _, _, p in nodes)}")

    print()
    passed = 0
    failed = 0
    skipped = 0

    for (inst, algo), exp_obj in sorted(expected.items()):
        cycle = load_cycle(inst, algo)
        if cycle is None:
            print(f"  SKIP  {inst:5s} {algo:10s} — brak pliku checker")
            skipped += 1
            continue

        checker_obj = checker_objective(instances[inst], cycle)
        ok = checker_obj == exp_obj
        status = "OK  " if ok else "FAIL"
        diff = f"  diff={checker_obj - exp_obj}" if not ok else ""
        print(f"  {status} {inst:5s} {algo:10s}  "
              f"ours={exp_obj:8d}  checker={checker_obj:8d}{diff}")

        if ok:
            passed += 1
        else:
            failed += 1

    print(f"\n{'=' * 55}")
    print(f"Passed: {passed}  Failed: {failed}  Skipped: {skipped}")
    if failed == 0 and skipped == 0:
        print("Wszystkie wyniki zgodne z Solution Checker!")
    return failed == 0


if __name__ == "__main__":
    print("Weryfikacja wyników vs Solution Checker\n")
    all_ok = check_all()
    raise SystemExit(0 if all_ok else 1)
